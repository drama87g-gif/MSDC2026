import os
import csv
import io
import functools
from datetime import datetime, timedelta
from flask import Flask, jsonify, request, send_from_directory, send_file, render_template_string
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy.orm import joinedload
from flask_migrate import Migrate
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
import openpyxl

# Configuration
app = Flask(__name__)

# Read database URL from env (default to sqlite for dev)
DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:////app/msdc.db')
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = False
app.config['UPLOAD_FOLDER'] = os.environ.get('UPLOAD_FOLDER', '/app/uploads')
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB max photo
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret')
app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY', app.config['SECRET_KEY'])
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(days=int(os.environ.get('JWT_EXP_DAYS', '1')))

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)
migrate = Migrate(app, db)
CORS(app)
jwt = JWTManager(app)

# Models
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(128), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(64), default='admin')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, nullable=True)
    username = db.Column(db.String(128), nullable=True)
    action = db.Column(db.String(128), nullable=False)
    model = db.Column(db.String(128), nullable=True)
    object_id = db.Column(db.String(128), nullable=True)
    details = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Patient(db.Model):
    __tablename__ = 'patients'
    __table_args__ = (
        db.Index('ix_patients_file_number', 'file_number'),
        db.Index('ix_patients_barcode', 'barcode'),
    )

    id = db.Column(db.Integer, primary_key=True)
    file_number = db.Column(db.String(32), unique=True, nullable=False)
    barcode = db.Column(db.String(64), unique=True, nullable=False)
    first_name = db.Column(db.String(128), nullable=False)
    last_name = db.Column(db.String(128), nullable=False)
    national_id = db.Column(db.String(64))
    nationality = db.Column(db.String(64))
    date_of_birth = db.Column(db.Date)
    gender = db.Column(db.String(16))
    address = db.Column(db.Text)
    phone = db.Column(db.String(32))
    email = db.Column(db.String(254))
    medical_history = db.Column(db.Text)
    allergies = db.Column(db.Text)
    photo_filename = db.Column(db.String(256))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'file_number': self.file_number,
            'barcode': self.barcode,
            'first_name': self.first_name,
            'last_name': self.last_name,
            'national_id': self.national_id,
            'nationality': self.nationality,
            'date_of_birth': self.date_of_birth.isoformat() if self.date_of_birth else None,
            'gender': self.gender,
            'address': self.address,
            'phone': self.phone,
            'email': self.email,
            'medical_history': self.medical_history,
            'allergies': self.allergies,
            'photo_url': f"/uploads/{self.photo_filename}" if self.photo_filename else None,
            'created_at': self.created_at.isoformat(),
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }

    def to_summary(self):
        return {
            'id': self.id,
            'file_number': self.file_number,
            'barcode': self.barcode,
            'first_name': self.first_name,
            'last_name': self.last_name,
            'phone': self.phone,
            'date_of_birth': self.date_of_birth.isoformat() if self.date_of_birth else None,
            'photo_url': f"/uploads/{self.photo_filename}" if self.photo_filename else None,
        }


class Medication(db.Model):
    __tablename__ = 'medications'
    __table_args__ = (
        db.Index('ix_medications_barcode', 'barcode'),
    )

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(256), nullable=False)
    generic_name = db.Column(db.String(256))
    barcode = db.Column(db.String(64), unique=True)
    dosage_form = db.Column(db.String(64))
    concentration = db.Column(db.String(64))
    quantity = db.Column(db.Integer, default=0)
    expiration_date = db.Column(db.Date)
    pharmaceutical_class = db.Column(db.String(128))
    manufacturer = db.Column(db.String(256))
    price = db.Column(db.Float)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'generic_name': self.generic_name,
            'barcode': self.barcode,
            'dosage_form': self.dosage_form,
            'concentration': self.concentration,
            'quantity': self.quantity,
            'expiration_date': self.expiration_date.isoformat() if self.expiration_date else None,
            'pharmaceutical_class': self.pharmaceutical_class,
            'manufacturer': self.manufacturer,
            'price': self.price,
            'created_at': self.created_at.isoformat()
        }


class Appointment(db.Model):
    __tablename__ = 'appointments'

    id = db.Column(db.Integer, primary_key=True)
    patient_id = db.Column(db.Integer, db.ForeignKey('patients.id'), nullable=False, index=True)
    appointment_type = db.Column(db.String(64), nullable=False)
    scheduled_date = db.Column(db.DateTime, nullable=False)
    status = db.Column(db.String(32), default='scheduled')
    queue_number = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    patient = db.relationship('Patient', backref='appointments')

    def to_dict(self):
        return {
            'id': self.id,
            'patient_id': self.patient_id,
            'patient_name': f"{self.patient.first_name} {self.patient.last_name}" if self.patient else None,
            'appointment_type': self.appointment_type,
            'scheduled_date': self.scheduled_date.isoformat() if self.scheduled_date else None,
            'status': self.status,
            'queue_number': self.queue_number,
            'created_at': self.created_at.isoformat()
        }


# Helpers
def paginate_query(query, schema_func=None):
    try:
        page = int(request.args.get('page', 1))
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get('per_page', 25))
    except ValueError:
        per_page = 25
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    items = [schema_func(i) if schema_func else i.to_dict() for i in pagination.items]
    return {
        'items': items,
        'page': pagination.page,
        'per_page': pagination.per_page,
        'total': pagination.total,
        'pages': pagination.pages
    }


# Simple audit logging helper
def log_action(user_identity, action, model=None, object_id=None, details=None):
    try:
        user_id = None
        username = None
        if user_identity:
            user_id = user_identity.get('id') if isinstance(user_identity, dict) else None
            username = user_identity.get('username') if isinstance(user_identity, dict) else None
        entry = AuditLog(user_id=user_id, username=username, action=action, model=model, object_id=str(object_id) if object_id else None, details=details)
        db.session.add(entry)
        db.session.commit()
    except Exception:
        # don't block main flow on audit errors
        db.session.rollback()


# Role check decorator
def role_required(*allowed_roles):
    def decorator(fn):
        @functools.wraps(fn)
        @jwt_required()
        def wrapper(*args, **kwargs):
            identity = get_jwt_identity()
            role = identity.get('role') if isinstance(identity, dict) else None
            if role not in allowed_roles and 'admin' not in allowed_roles and role != 'admin' and 'admin' not in allowed_roles:
                return jsonify({'msg': 'forbidden'}), 403
            return fn(*args, **kwargs)
        return wrapper
    return decorator


# Authentication endpoints
@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    username = data.get('username')
    password = data.get('password')
    if not username or not password:
        return jsonify({'msg': 'username and password required'}), 400
    user = User.query.filter_by(username=username).first()
    if not user or not user.check_password(password):
        return jsonify({'msg': 'Bad username or password'}), 401
    access_token = create_access_token(identity={'id': user.id, 'username': user.username, 'role': user.role})
    return jsonify({'access_token': access_token}), 200


# API Routes
@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({
        'status': 'ok',
        'system': 'MSDC Hospital Management System',
        'version': '1.0.0',
        'timestamp': datetime.utcnow().isoformat()
    })


# Patients - list with pagination, create, import, export
@app.route('/api/patients', methods=['GET', 'POST'])
def patients():
    if request.method == 'POST':
        # protected: only admin or admission
        return _create_patient()

    # GET - paginated summary
    query = Patient.query.order_by(Patient.id.desc())
    return jsonify(paginate_query(query, schema_func=lambda p: p.to_summary()))


@role_required('admin', 'admission')
def _create_patient():
    data = request.get_json() or {}
    # Parse date_of_birth if provided
    dob = None
    if data.get('date_of_birth'):
        try:
            dob = datetime.fromisoformat(data.get('date_of_birth')).date()
        except Exception:
            dob = None
    patient = Patient(
        file_number=data.get('file_number') or '',
        barcode=data.get('barcode') or '',
        first_name=data.get('first_name') or '',
        last_name=data.get('last_name') or '',
        national_id=data.get('national_id'),
        nationality=data.get('nationality'),
        date_of_birth=dob,
        gender=data.get('gender'),
        address=data.get('address'),
        phone=data.get('phone'),
        email=data.get('email'),
        medical_history=data.get('medical_history'),
        allergies=data.get('allergies')
    )
    db.session.add(patient)
    db.session.commit()
    # log
    log_action(get_jwt_identity(), 'create_patient', model='Patient', object_id=patient.id)
    return jsonify(patient.to_dict()), 201


@app.route('/api/patients/<int:patient_id>', methods=['GET', 'PUT', 'DELETE'])
@jwt_required(optional=True)
def patient_detail(patient_id):
    patient = Patient.query.get_or_404(patient_id)

    if request.method == 'GET':
        return jsonify(patient.to_dict())

    identity = get_jwt_identity()
    # PUT requires admin or admission
    if request.method == 'PUT':
        if not identity or identity.get('role') not in ('admin', 'admission'):
            return jsonify({'msg': 'forbidden'}), 403
        data = request.get_json() or {}
        for key, value in data.items():
            if hasattr(patient, key):
                # Basic conversion for date
                if key == 'date_of_birth' and value:
                    try:
                        setattr(patient, key, datetime.fromisoformat(value).date())
                    except Exception:
                        pass
                else:
                    setattr(patient, key, value)
        db.session.commit()
        log_action(identity, 'update_patient', model='Patient', object_id=patient.id)
        return jsonify(patient.to_dict())

    # DELETE
    if request.method == 'DELETE':
        if not identity or identity.get('role') not in ('admin',):
            return jsonify({'msg': 'forbidden'}), 403
        # remove photo file if exists
        if patient.photo_filename:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], patient.photo_filename))
            except Exception:
                pass
        db.session.delete(patient)
        db.session.commit()
        log_action(identity, 'delete_patient', model='Patient', object_id=patient.id)
        return '', 204


# Photo upload
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/api/patients/<int:patient_id>/photo', methods=['POST'])
@role_required('admin', 'admission')
def upload_patient_photo(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filename = f"patient_{patient.id}_{int(datetime.utcnow().timestamp())}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        patient.photo_filename = filename
        db.session.commit()
        log_action(get_jwt_identity(), 'upload_photo', model='Patient', object_id=patient.id)
        return jsonify({'photo_url': f"/uploads/{filename}"}), 201
    return jsonify({'error': 'Invalid file type'}), 400


@app.route('/uploads/<path:filename>', methods=['GET'])
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


# Export patients to CSV/XLSX
@app.route('/api/patients/export', methods=['GET'])
@role_required('admin', 'admission', 'reception')
def export_patients():
    fmt = request.args.get('format', 'csv')
    patients_q = Patient.query.order_by(Patient.id)

    if fmt == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['id', 'file_number', 'barcode', 'first_name', 'last_name', 'national_id', 'nationality', 'date_of_birth', 'phone', 'address', 'medical_history', 'allergies']
        ws.append(headers)
        for p in patients_q:
            ws.append([
                p.id,
                p.file_number,
                p.barcode,
                p.first_name,
                p.last_name,
                p.national_id,
                p.nationality,
                p.date_of_birth.isoformat() if p.date_of_birth else '',
                p.phone,
                p.address,
                p.medical_history,
                p.allergies
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        log_action(get_jwt_identity(), 'export_patients', details='format=xlsx')
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='patients_export.xlsx')

    si = io.StringIO()
    cw = csv.writer(si)
    headers = ['id', 'file_number', 'barcode', 'first_name', 'last_name', 'national_id', 'nationality', 'date_of_birth', 'phone', 'address', 'medical_history', 'allergies']
    cw.writerow(headers)
    for p in patients_q:
        cw.writerow([
            p.id,
            p.file_number,
            p.barcode,
            p.first_name,
            p.last_name,
            p.national_id,
            p.nationality,
            p.date_of_birth.isoformat() if p.date_of_birth else '',
            p.phone,
            p.address,
            p.medical_history,
            p.allergies
        ])
    output = io.BytesIO()
    output.write(si.getvalue().encode('utf-8'))
    output.seek(0)
    log_action(get_jwt_identity(), 'export_patients', details='format=csv')
    return send_file(output, mimetype='text/csv', as_attachment=True, download_name='patients_export.csv')


# Import patients from CSV/XLSX
@app.route('/api/patients/import', methods=['POST'])
@role_required('admin', 'admission')
def import_patients():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    filename = file.filename.lower()
    created = 0

    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [h for h in rows[0]]
        for row in rows[1:]:
            rowd = dict(zip(headers, row))
            dob = None
            if rowd.get('date_of_birth'):
                try:
                    dob = datetime.fromisoformat(str(rowd.get('date_of_birth'))).date()
                except Exception:
                    dob = None
            patient = Patient(
                file_number=str(rowd.get('file_number') or ''),
                barcode=str(rowd.get('barcode') or ''),
                first_name=str(rowd.get('first_name') or ''),
                last_name=str(rowd.get('last_name') or ''),
                national_id=str(rowd.get('national_id') or None),
                nationality=str(rowd.get('nationality') or None),
                date_of_birth=dob,
                phone=str(rowd.get('phone') or None),
                address=str(rowd.get('address') or None),
                medical_history=str(rowd.get('medical_history') or None),
                allergies=str(rowd.get('allergies') or None)
            )
            db.session.add(patient)
            created += 1
        db.session.commit()
        log_action(get_jwt_identity(), 'import_patients', details='format=xlsx')
        return jsonify({'created': created}), 201

    # else assume CSV
    stream = io.StringIO(file.stream.read().decode('utf-8'))
    reader = csv.DictReader(stream)
    for row in reader:
        dob = None
        if row.get('date_of_birth'):
            try:
                dob = datetime.fromisoformat(row.get('date_of_birth')).date()
            except Exception:
                dob = None
        patient = Patient(
            file_number=row.get('file_number') or '',
            barcode=row.get('barcode') or '',
            first_name=row.get('first_name') or '',
            last_name=row.get('last_name') or '',
            national_id=row.get('national_id'),
            nationality=row.get('nationality'),
            date_of_birth=dob,
            phone=row.get('phone'),
            address=row.get('address'),
            medical_history=row.get('medical_history'),
            allergies=row.get('allergies')
        )
        db.session.add(patient)
        created += 1
    db.session.commit()
    log_action(get_jwt_identity(), 'import_patients', details='format=csv')
    return jsonify({'created': created}), 201


# Printable patient card (HTML)
CARD_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Patient Card</title>
  <style>
    body { font-family: Arial, sans-serif; }
    .card { width: 350px; height: 200px; border: 1px solid #333; padding: 10px; }
    .photo { float: right; width: 90px; height: 90px; object-fit: cover; border: 1px solid #ccc; }
    .row { margin-bottom: 6px; }
    .label { font-weight: bold; }
  </style>
</head>
<body>
  <div class="card">
    {% if photo_url %}
      <img src="{{ photo_url }}" class="photo" />
    {% endif %}
    <div class="row"><span class="label">Name:</span> {{ first_name }} {{ last_name }}</div>
    <div class="row"><span class="label">National ID:</span> {{ national_id }}</div>
    <div class="row"><span class="label">File #:</span> {{ file_number }}</div>
    <div class="row"><span class="label">Nationality:</span> {{ nationality }}</div>
    <div class="row"><span class="label">Phone:</span> {{ phone }}</div>
    <div class="row"><span class="label">Address:</span> {{ address }}</div>
    <div class="row"><span class="label">Allergies:</span> {{ allergies }}</div>
    <div class="row"><span class="label">Barcode:</span> {{ barcode }}</div>
  </div>
</body>
</html>
"""


@app.route('/api/patients/<int:patient_id>/card', methods=['GET'])
def patient_card(patient_id):
    patient = Patient.query.get_or_404(patient_id)
    return render_template_string(CARD_TEMPLATE, **patient.to_dict())


# Medications - paginated
@app.route('/api/medications', methods=['GET', 'POST'])
def medications():
    if request.method == 'POST':
        # protected: only admin or pharmacy
        identity = None
        try:
            identity = get_jwt_identity()
        except Exception:
            pass
        if not identity or identity.get('role') not in ('admin', 'pharmacy'):
            return jsonify({'msg': 'forbidden'}), 403
        data = request.get_json() or {}
        exp_date = None
        if data.get('expiration_date'):
            try:
                exp_date = datetime.fromisoformat(data.get('expiration_date')).date()
            except Exception:
                exp_date = None
        medication = Medication(
            name=data.get('name'),
            generic_name=data.get('generic_name'),
            barcode=data.get('barcode'),
            dosage_form=data.get('dosage_form'),
            concentration=data.get('concentration'),
            quantity=int(data.get('quantity', 0)),
            expiration_date=exp_date,
            pharmaceutical_class=data.get('pharmaceutical_class'),
            manufacturer=data.get('manufacturer'),
            price=data.get('price')
        )
        db.session.add(medication)
        db.session.commit()
        log_action(identity, 'create_medication', model='Medication', object_id=medication.id)
        return jsonify(medication.to_dict()), 201

    query = Medication.query.order_by(Medication.id.desc())
    return jsonify(paginate_query(query, schema_func=lambda m: m.to_dict()))


# Appointments - paginated with eager-loaded patient
@app.route('/api/appointments', methods=['GET', 'POST'])
def appointments():
    if request.method == 'POST':
        # protected: reception or admin
        identity = None
        try:
            identity = get_jwt_identity()
        except Exception:
            pass
        if not identity or identity.get('role') not in ('admin', 'reception'):
            return jsonify({'msg': 'forbidden'}), 403
        data = request.get_json() or {}
        sched = None
        if data.get('scheduled_date'):
            try:
                sched = datetime.fromisoformat(data.get('scheduled_date'))
            except Exception:
                sched = None
        appointment = Appointment(
            patient_id=data.get('patient_id'),
            appointment_type=data.get('appointment_type'),
            scheduled_date=sched,
            queue_number=data.get('queue_number')
        )
        db.session.add(appointment)
        db.session.commit()
        log_action(identity, 'create_appointment', model='Appointment', object_id=appointment.id)
        return jsonify(appointment.to_dict()), 201

    query = Appointment.query.options(joinedload(Appointment.patient)).order_by(Appointment.scheduled_date.desc())
    return jsonify(paginate_query(query, schema_func=lambda a: a.to_dict()))


# Admin seed helper
def seed_admin():
    admin_username = os.environ.get('SEED_ADMIN_USER', 'admin')
    admin_password = os.environ.get('SEED_ADMIN_PASS', 'AdminPass123')
    if User.query.filter_by(username=admin_username).first():
        return
    pw_hash = generate_password_hash(admin_password)
    user = User(username=admin_username, password_hash=pw_hash, role='admin')
    db.session.add(user)
    db.session.commit()
    print('Seeded admin user:', admin_username)


# Serve app
if __name__ == '__main__':
    # If INIT_DB environment var is set to "1", create tables at startup (dev only) and seed admin
    if os.environ.get('INIT_DB') == '1':
        with app.app_context():
            db.create_all()
            try:
                seed_admin()
            except Exception:
                pass
    # Use gunicorn in production via Docker CMD; here fallback to Flask dev server for local dev
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=os.environ.get('FLASK_DEBUG', '0') == '1')
